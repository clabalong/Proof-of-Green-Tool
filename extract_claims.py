"""
GreenLens — Environmental Claim Extractor
==========================================
Reads scraped company JSON files, extracts environmental/sustainability claims
using the Claude API, independently verifies them using a separate OpenAI
model, and writes results to an Excel file.

Usage:
    python extract_claims.py                          # process all JSONs in ./data/
    python extract_claims.py --input ./data/          # same, explicit folder
    python extract_claims.py --input ballymaloe.json  # single file
    python extract_claims.py --input ./data/ --output results.xlsx

Programmatic usage (e.g. from pipelineV1.py, chained straight after Stage 1):
    from extract_claims import process_scrape_result
    claims_result = process_scrape_result(anthropic_client, openai_client, scrape_result_dict)

Requirements:
    pip install anthropic openai openpyxl python-dotenv

API keys:
    export ANTHROPIC_API_KEY="sk-ant-..."
    export OPENAI_API_KEY="sk-..."
"""

import os
import json
import argparse
import sys
from pathlib import Path
from datetime import datetime

import anthropic
import openai
import json_repair
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from dotenv import load_dotenv
load_dotenv()


# ── Prompt ────────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """You are an environmental claims extractor for an EU greenwashing compliance research project.

Your task: read the text from a company web page and extract EVERY environmental or sustainability-related claim — explicit or implicit, strong or vague.

Include:
- Direct environmental claims ("100% recycled packaging", "carbon neutral")
- Certification mentions ("Origin Green member", "B Corp certified")
- Sourcing / supply chain claims ("locally sourced", "free-range")
- Biodiversity / nature claims ("pollinator-friendly", "supports birdwatch")
- Waste / circular economy claims ("zero waste", "redistributes surplus food")
- Vague green language ("sustainable", "eco-friendly", "natural", "responsible")
- Quantified claims ("reduced CO₂ by 43 tonnes", "428 kg of virgin plastic replaced")

Do NOT include:
- Pure product descriptions with no environmental angle
- Navigation links, cookie notices, contact details

Extract complete claim-bearing passages rather than isolated keywords, section headings, article titles, podcast titles, or navigation text.
When a claim depends on surrounding text, include enough of the original sentence or adjacent sentence to preserve its subject, scope, and qualification.
Do not extract a standalone word such as "sustainability", "organic", or "green" unless it forms part of a product name or phrase that itself communicates an environmental characteristic to consumers.

Return ONLY a valid JSON array. Each object must have these exact fields:
{
  "text": "verbatim claim as it appears on the page",
  "language": "ISO 639-1 code, e.g. en / fr / de / nl",
  "english_translation": "English translation (identical to text if already English)",
  "category": "one of: carbon | biodiversity | packaging | water | sourcing | certification | waste | general",
  "confidence": 0.85
}

If the page contains no environmental claims, return [].
No preamble, no markdown fences, no trailing text — just the JSON array."""


# ── Verification (independent second model) ────────────────────────────────

VERIFICATION_MODEL = "gpt-5.4-mini"  # OpenAI's current flagship as of mid-2026;
                                 # use "gpt-5.4-mini" instead for a cheaper pass

VERIFICATION_SYSTEM_PROMPT = """You are an independent environmental-claim validation auditor.

You will receive:
1. The raw text from a company webpage.
2. Environmental claims extracted by another AI model.

For each extracted item, independently determine whether it is a valid environmental marketing claim suitable for ECGT compliance classification.

============================================================
STAGE 1 — TEXTUAL FAITHFULNESS
============================================================
Determine whether the extracted claim appears in the supplied page text.

The claim does not need to be character-for-character identical, but it must be a faithful, near-exact representation of the page text.

Set text_verified = false when:
- The claim is not present in the page text.
- The extractor invented information.
- The extractor exaggerated the original wording.
- The extractor combined separate passages in a misleading way.
- Important qualifications were removed.
- The English translation changes the claim's meaning.

============================================================
STAGE 2 — ENVIRONMENTAL RELEVANCE
============================================================
Determine whether the item asserts or implies an environmental, sustainability, circularity, sourcing, nature, waste, emissions, certification, farming-practice, or resource-use characteristic.

Valid environmental claims include:
- Specific environmental performance claims.
- Generic environmental language.
- Certification or environmental label claims.
- Organic, regenerative, biodynamic or permaculture claims.
- Local or responsible sourcing claims with an environmental implication.
- Carbon, emissions, biodiversity, water, waste or packaging claims.
- Environmental aspirations or future commitments.
- Statements intended to influence consumer perceptions of environmental performance.

Do not reject a claim merely because it is vague, unsubstantiated or potentially misleading. Those matters belong to the later ECGT classification stage.

============================================================
STAGE 3 — COMPANY ATTRIBUTION
============================================================
Determine whether the page presents the statement as a claim made by, adopted by, or clearly attributable to the company or its products.

Set company_attributed = false when the text is only:
- A statement made by an unrelated third party.
- A quotation that the company does not adopt or endorse.
- A news headline about another organisation.
- A guest's opinion in a podcast or interview.
- A general educational statement not connected to the company's own products, operations, commitments or services.

A statement may still be company-attributed when:
- It appears in the company's own product description.
- It describes the company's farm, packaging, sourcing or operations.
- It expresses the company's mission, goal or commitment.
- The company republishes the statement as part of its marketing message.

============================================================
STAGE 4 — COMPLETE, ASSESSABLE CLAIM
============================================================
Determine whether the extracted text contains enough meaning to be assessed.

Set complete_claim = false when it is merely:
- Navigation text.
- A button such as "Read more".
- A standalone page-section label such as "Sustainability".
- A product or article title with no environmental assertion.
- A podcast episode title that does not itself assert something about the company or its products.
- An incomplete sentence fragment whose meaning depends on missing text.
- Contact details, cookie text or administrative content.

Important distinctions:

"Farm Veg Box"
→ Not a complete environmental claim.

"Organic Farm Veg Box"
→ May be a valid environmental/product claim because "organic" asserts a regulated product characteristic.

"Sustainability"
→ Not a complete claim.

"We make sustainability central to every decision"
→ Valid environmental claim, although possibly vague.

"Episode 42: Climate Change"
→ Normally not a company environmental claim.

"We reduced packaging weight by 20% in 2025"
→ Valid environmental claim.

"Our farm sequesters carbon every year"
→ Valid environmental claim. Do not assess whether it is proven at this stage.

============================================================
STAGE 5 — CATEGORY CHECK
============================================================
Review the assigned category.

Allowed categories:
carbon | biodiversity | packaging | water | sourcing | certification | waste | general

Return the most appropriate category. Do not mark an otherwise valid claim as invalid merely because the original category was wrong.

============================================================
FINAL DECISION
============================================================
Set valid_claim = true only when all of the following are true:
- text_verified is true
- environmental_relevance is true
- company_attributed is true
- complete_claim is true

Otherwise set valid_claim = false.

Return one result for every supplied item, in exactly the same order.

Return ONLY a valid JSON array with these exact fields:
{
  "id": 0,
  "text_verified": true,
  "environmental_relevance": true,
  "company_attributed": true,
  "complete_claim": true,
  "valid_claim": true,
  "corrected_category": "packaging",
  "exclusion_reason": null,
  "notes": "Brief explanation of the decision."
}

For exclusion_reason, use exactly one of:
- null
- "NOT_FOUND_IN_TEXT"
- "PARAPHRASED_OR_EXAGGERATED"
- "NOT_ENVIRONMENTAL"
- "NOT_COMPANY_ATTRIBUTED"
- "TITLE_OR_NAVIGATION"
- "INCOMPLETE_FRAGMENT"
- "OTHER"

No preamble.
No markdown fences.
No trailing text.
Only the JSON array.
""".strip()


def verify_claims_with_gpt(openai_client, page_text: str, claims: list[dict], verbose: bool = True) -> list[dict]:
    """
    Independently verifies a list of already-extracted claims using a
    SEPARATE model from a different provider, rather than having Claude
    check its own work. Returns the same claims list with "gpt_verified"
    (bool or None on failure) and "gpt_notes" (str) added to each claim.
    """
    if not claims:
        return claims

    numbered = [
        {"id": i, "text": c.get("text", ""), "category": c.get("category", "general")}
        for i, c in enumerate(claims)
    ]
    user_msg = (
        f"Page text:\n{page_text[:8000]}\n\n"
        f"Claims to verify:\n{json.dumps(numbered, ensure_ascii=False, indent=2)}"
    )

    try:
        response = openai_client.chat.completions.create(
            model=VERIFICATION_MODEL,
            messages=[
                {"role": "system", "content": VERIFICATION_SYSTEM_PROMPT},
                {"role": "user", "content": user_msg},
            ],
        )
        raw = response.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
            raw = raw.strip()

        try:
            results = json.loads(raw)
        except json.JSONDecodeError:
            results = json_repair.loads(raw)

        results_by_id = {r["id"]: r for r in results if isinstance(r, dict) and "id" in r}
        allowed_categories = {
            "carbon", "biodiversity", "packaging", "water",
            "sourcing", "certification", "waste", "general"
        }

        for i, claim in enumerate(claims):
            result = results_by_id.get(i)

            if result:
                claim["gpt_text_verified"] = result.get("text_verified")
                claim["gpt_environmental_relevance"] = result.get("environmental_relevance")
                claim["gpt_company_attributed"] = result.get("company_attributed")
                claim["gpt_complete_claim"] = result.get("complete_claim")
                claim["gpt_valid_claim"] = result.get("valid_claim")
                claim["gpt_corrected_category"] = result.get(
                    "corrected_category", claim.get("category", "general")
                )
                claim["gpt_exclusion_reason"] = result.get("exclusion_reason")
                claim["gpt_notes"] = result.get("notes", "")

                # Backward compatibility with the existing workbook field.
                claim["gpt_verified"] = result.get("valid_claim")

                # Apply a corrected category only to valid claims.
                corrected = result.get("corrected_category")
                if result.get("valid_claim") is True and corrected in allowed_categories:
                    claim["category"] = corrected
            else:
                claim["gpt_text_verified"] = None
                claim["gpt_environmental_relevance"] = None
                claim["gpt_company_attributed"] = None
                claim["gpt_complete_claim"] = None
                claim["gpt_valid_claim"] = None
                claim["gpt_corrected_category"] = ""
                claim["gpt_exclusion_reason"] = "OTHER"
                claim["gpt_verified"] = None
                claim["gpt_notes"] = "No verification result returned for this claim"

        if verbose:
            verified_count = sum(1 for c in claims if c.get("gpt_valid_claim") is True)
            print(f"[GPT validation: {verified_count}/{len(claims)} valid claims]", end=" ")

    except Exception as e:
        if verbose:
            print(f"[GPT verification FAILED: {type(e).__name__} — {e}]", end=" ")
        for c in claims:
            c["gpt_text_verified"] = None
            c["gpt_environmental_relevance"] = None
            c["gpt_company_attributed"] = None
            c["gpt_complete_claim"] = None
            c["gpt_valid_claim"] = None
            c["gpt_corrected_category"] = ""
            c["gpt_exclusion_reason"] = "OTHER"
            c["gpt_verified"] = None
            c["gpt_notes"] = f"Verification pass failed: {e}"

    return claims


# ── Core extraction ────────────────────────────────────────────────────────────

def extract_claims_from_page(client, company_name: str, page_name: str, page_url: str, page_text: str) -> list[dict]:
    """Call Claude API for one page and return a list of claim dicts."""
    user_msg = f"Company: {company_name}\nPage: {page_name}\nURL: {page_url}\n\nPage text:\n{page_text[:8000]}"

    try:
        message = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=10000,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": user_msg}],
        )
        raw = message.content[0].text.strip()

        # Strip accidental markdown fences
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
            raw = raw.strip()

        claims = json.loads(raw)
        if not isinstance(claims, list):
            return []

        # Attach page metadata to each claim
        for c in claims:
            c["page"] = page_name
            c["page_url"] = page_url

        return claims

    except json.JSONDecodeError as e:
        # Common cause: a verbatim claim contains an actual quote
        # character (a testimonial, guillemets normalized to straight
        # quotes, etc.) that the model didn't escape, breaking strict
        # JSON syntax. Try a permissive repair before giving up —
        # recovers the claims instead of silently losing the whole page.
        try:
            claims = json_repair.loads(raw)
            if isinstance(claims, list) and claims:
                print(f"    [WARN] JSON parse error on '{page_name}' ({e}) — "
                      f"recovered {len(claims)} claims via json_repair")
                for c in claims:
                    c["page"] = page_name
                    c["page_url"] = page_url
                return claims
        except Exception:
            pass

        print(f"    [WARN] JSON parse error on '{page_name}': {e}")
        # Save the raw response so we can inspect exactly what broke,
        # instead of silently discarding it.
        debug_dir = Path("debug_failed_json")
        debug_dir.mkdir(exist_ok=True)
        debug_path = debug_dir / f"{company_name.lower().replace(' ', '_')}_{page_name.lower()}.txt"
        try:
            with open(debug_path, "w", encoding="utf-8") as f:
                f.write(raw)
            print(f"    [WARN] Raw response saved to: {debug_path}")
        except Exception:
            pass
        return []
    except anthropic.APIError as e:
        print(f"    [ERROR] API error on '{page_name}': {e}")
        return []


def process_scrape_result(anthropic_client, openai_client, data: dict, source_file: str = "", verbose: bool = True) -> dict:
    """
    Process one company's scrape result — takes the SAME dict shape that
    run_single_scrape() / scrape_sme_website() returns, whether it came
    from a freshly-scraped in-memory result or was loaded from JSON.

    This is the function pipelineV1.py calls directly, so Stage 1 output
    can flow straight into Stage 2 without a mandatory disk round-trip.

    Args:
        anthropic_client: an anthropic.Anthropic client instance (extraction)
        openai_client: an openai.OpenAI client instance (independent verification)
        data: dict with keys company_name, base_url, scraped_at, pages
        source_file: optional label for the Excel "Source File" column
        verbose: whether to print progress to stdout

    Returns:
        dict with company_name, base_url, scraped_at, source_file, claims
    """
    company_name = data.get("company_name", "Unknown Company")
    base_url = data.get("base_url", "")
    scraped_at = data.get("scraped_at", "")
    pages = data.get("pages", {})

    if verbose:
        print(f"\n{'─'*60}")
        print(f"  Company : {company_name}")
        if source_file:
            print(f"  File    : {source_file}")
        print(f"  Pages   : {list(pages.keys())}")

    all_claims = []
    for page_name, page_data in pages.items():
        page_text = page_data.get("text", "").strip()
        page_url = page_data.get("url", "")

        if len(page_text) < 50:
            if verbose:
                print(f"    [{page_name}] skipped — too short")
            continue

        if verbose:
            print(f"    [{page_name}] extracting...", end=" ", flush=True)
        claims = extract_claims_from_page(anthropic_client, company_name, page_name, page_url, page_text)
        if verbose:
            print(f"{len(claims)} claims found", end=" ")

        if claims:
            claims = verify_claims_with_gpt(openai_client, page_text, claims, verbose=verbose)

        if verbose:
            print()

        all_claims.extend(claims)

    if verbose:
        print(f"  Total   : {len(all_claims)} claims")

    return {
        "company_name": company_name,
        "base_url": base_url,
        "scraped_at": scraped_at,
        "source_file": source_file,
        "claims": all_claims,
    }


def process_company_file(anthropic_client, openai_client, json_path: Path) -> dict:
    """Process one company JSON file and return extracted claims + metadata.
    Thin wrapper around process_scrape_result() for the file-based CLI/batch path."""
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    return process_scrape_result(anthropic_client, openai_client, data, source_file=json_path.name)


# ── Excel output ───────────────────────────────────────────────────────────────

CATEGORY_COLORS = {
    "carbon":        "D6EAF8",
    "biodiversity":  "D5F5E3",
    "packaging":     "FAD7A0",
    "water":         "AED6F1",
    "sourcing":      "E8DAEF",
    "certification": "FDEBD0",
    "waste":         "D5DBDB",
    "general":       "F9EBEA",
}

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def write_excel(results: list[dict], output_path: Path):
    wb = Workbook()

    # ── Sheet 1: All claims ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "All Claims"

    headers = [
        "Company", "Page", "Category", "Verbatim Claim",
        "English Translation", "Language", "Confidence",
        "GPT Valid Claim", "Text Verified", "Environmental Relevance",
        "Company Attributed", "Complete Claim", "Exclusion Reason", "GPT Notes",
        "Page URL", "Source File",
        "EMAS Verified", "EU Organic Verified", "B Corp Verified", "Bord Bia Verified",
        "Biopartenaire Verified", "BioED Verified",
        "ECGT Label", "ECGT Rule Triggered", "ECGT Explanation", "ECGT Review Flag"
    ]
    col_widths = [22, 16, 14, 60, 60, 10, 12, 14, 14, 20, 18, 14, 22, 45, 45, 25, 14, 16, 14, 15, 18, 14,
                  12, 32, 55, 14]

    # Header row
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    row_idx = 2
    for res in results:
        for c in res["claims"]:
            conf = c.get("confidence", 0)
            cat = c.get("category", "general").lower()
            fill_color = CATEGORY_COLORS.get(cat, "FFFFFF")
            row_fill = PatternFill("solid", fgColor=fill_color)

            def _yn(key):
                val = c.get(key)
                if val is None:
                    return ""  # cert stage / verification wasn't run or merged for this claim
                return "Yes" if val else "No"

            values = [
                res["company_name"],
                c.get("page", ""),
                c.get("category", "general"),
                c.get("text", ""),
                c.get("english_translation", c.get("text", "")),
                c.get("language", "en").upper(),
                round(conf, 2),
                _yn("gpt_valid_claim"),
                _yn("gpt_text_verified"),
                _yn("gpt_environmental_relevance"),
                _yn("gpt_company_attributed"),
                _yn("gpt_complete_claim"),
                c.get("gpt_exclusion_reason", ""),
                c.get("gpt_notes", ""),
                c.get("page_url", ""),
                res["source_file"],
                _yn("emas_verified"),
                _yn("eu_organic_verified"),
                _yn("bcorp_verified"),
                _yn("bordbia_verified"),
                _yn("biopartenaire_verified"),
                _yn("bioed_verified"),
                c.get("ecgt_label", ""),
                c.get("ecgt_rule_triggered", ""),
                c.get("ecgt_explanation", ""),
                _yn("ecgt_review_flag"),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = row_fill
                cell.border = thin_border()
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if col_idx in (4, 5):  # claim text columns
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            ws.row_dimensions[row_idx].height = 40
            row_idx += 1

    # ── Sheet 2: Summary per company ──────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    sum_headers = ["Company", "Source File", "Base URL", "Scraped At",
                   "Total Claims", "carbon", "biodiversity", "packaging",
                   "water", "sourcing", "certification", "waste", "general",
                   "Languages Detected"]
    sum_widths =  [22, 25, 40, 22, 14, 10, 14, 12, 8, 10, 15, 8, 10, 25]

    for col_idx, (h, w) in enumerate(zip(sum_headers, sum_widths), 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
    ws2.row_dimensions[1].height = 28
    ws2.freeze_panes = "A2"

    categories = ["carbon", "biodiversity", "packaging", "water",
                  "sourcing", "certification", "waste", "general"]

    for row_idx, res in enumerate(results, 2):
        claims = res["claims"]
        cat_counts = {cat: sum(1 for c in claims if c.get("category", "general") == cat) for cat in categories}
        langs = ", ".join(sorted({c.get("language", "en").upper() for c in claims})) or "EN"

        row_data = [
            res["company_name"], res["source_file"], res["base_url"], res["scraped_at"],
            len(claims),
            *[cat_counts[cat] for cat in categories],
            langs,
        ]
        alt_fill = PatternFill("solid", fgColor="F2F3F4" if row_idx % 2 == 0 else "FFFFFF")
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = alt_fill
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center")

    wb.save(output_path)
    print(f"\n  Saved → {output_path}")


# ── CLI ────────────────────────────────────────────────────────────────────────

def collect_json_files(input_path: str) -> list[Path]:
    p = Path(input_path)
    if p.is_file():
        return [p]
    if p.is_dir():
        files = sorted(p.glob("*.json"))
        if not files:
            print(f"[ERROR] No JSON files found in {p}")
            sys.exit(1)
        return files
    print(f"[ERROR] Path not found: {p}")
    sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description="GreenLens — Environmental Claim Extractor")
    parser.add_argument("--input",  default='/Users/tunaerdem/Desktop/Scraped Data/Scraped companies/Tuna_extract',       help="Path to a JSON file or folder of JSONs (default: ./data)")
    parser.add_argument("--output", default="",             help="Output Excel path (default: greenlens_claims_<timestamp>.xlsx)")
    args = parser.parse_args()

    anthropic_key = os.environ.get("ANTHROPIC_API_KEY")
    if not anthropic_key:
        print("[ERROR] ANTHROPIC_API_KEY environment variable not set.")
        sys.exit(1)

    openai_key = os.environ.get("OPENAI_API_KEY")
    if not openai_key:
        print("[ERROR] OPENAI_API_KEY environment variable not set.")
        sys.exit(1)

    anthropic_client = anthropic.Anthropic(api_key=anthropic_key)
    openai_client = openai.OpenAI(api_key=openai_key)

    json_files = collect_json_files(args.input)
    print(f"Found {len(json_files)} JSON file(s) to process.")

    results = []
    for path in json_files:
        result = process_company_file(anthropic_client, openai_client, path)
        results.append(result)

    if not any(r["claims"] for r in results):
        print("\n[WARN] No claims were extracted across all files.")
        sys.exit(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = Path(args.output) if args.output else Path(f"greenlens_claims_{timestamp}.xlsx")
    write_excel(results, output_path)

    total_claims = sum(len(r["claims"]) for r in results)
    print(f"\n{'='*60}")
    print(f"  Done. {total_claims} claims extracted from {len(json_files)} company file(s).")
    print(f"  Output: {output_path}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()