"""
GreenLens — Environmental Claim Extractor (v2 — no second-model verification)
================================================================================
Reads scraped company JSON files, extracts environmental/sustainability claims
using the Claude API, and flags each claim as confidence-verified based on a
threshold applied to the extraction model's own "confidence" score — no
second (GPT) model call. Writes results to an Excel file.

Difference from the main version (extract_claims_updated.py):
    The main version sends every extracted claim to a separate OpenAI model
    for an independent multi-stage validation pass (text faithfulness,
    environmental relevance, company attribution, completeness, category
    check). This version removes that call entirely. Instead, a claim is
    marked verified if its own extraction "confidence" score meets
    CONFIDENCE_THRESHOLD (0.8 by default). This is faster and needs no
    OPENAI_API_KEY, at the cost of no longer being an independent
    cross-check — it's just thresholding the extractor's self-reported
    confidence. Everything else (extraction prompt, JSON-repair fallback,
    certification checks, news check) is identical to the main pipeline.

Usage:
    python extract_claimsNO_GPT_Check.py                          # process all JSONs in ./data/
    python extract_claimsNO_GPT_Check.py --input ./data/          # same, explicit folder
    python extract_claimsNO_GPT_Check.py --input ballymaloe.json  # single file
    python extract_claimsNO_GPT_Check.py --input ./data/ --output results.xlsx
    python extract_claimsNO_GPT_Check.py --threshold 0.7          # override the 0.8 default

Programmatic usage (e.g. from pipelineNO_GPT_Check.py, chained straight after Stage 1):
    from extract_claimsNO_GPT_Check import process_scrape_result
    claims_result = process_scrape_result(anthropic_client, scrape_result_dict)

Requirements:
    pip install anthropic openpyxl python-dotenv json_repair

API keys:
    export ANTHROPIC_API_KEY="sk-ant-..."
"""

import os
import json
import argparse
import sys
from pathlib import Path
from datetime import datetime

import anthropic
import json_repair
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from dotenv import load_dotenv
load_dotenv()


# ── Prompt ────────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = """You are an environmental claims extractor for an EU greenwashing compliance research project.

Your task: read the text from a company web page and extract EVERY environmental or sustainability-related claim — explicit or implicit, strong or vague.

Include:
- Direct environmental claims ("100% recycled packaging", "carbon neutral")
- Certification mentions ("Origin Green member", "B Corp certified")
- Sourcing / supply chain claims ("locally sourced", "free-range")
- Biodiversity / nature claims ("pollinator-friendly", "supports birdwatch")
- Waste / circular economy claims ("zero waste", "redistributes surplus food")
- Vague green language ("sustainable", "eco-friendly", "natural", "responsible")
- Quantified claims ("reduced CO₂ by 43 tonnes", "428 kg of virgin plastic replaced")

Do NOT include:
- Pure product descriptions with no environmental angle
- Navigation links, cookie notices, contact details

Extract complete claim-bearing passages rather than isolated keywords, section headings, article titles, podcast titles, or navigation text.
When a claim depends on surrounding text, include enough of the original sentence or adjacent sentence to preserve its subject, scope, and qualification.
Do not extract a standalone word such as "sustainability", "organic", or "green" unless it forms part of a product name or phrase that itself communicates an environmental characteristic to consumers.

Return ONLY a valid JSON array. Each object must have these exact fields:
{
  "text": "verbatim claim as it appears on the page",
  "language": "ISO 639-1 code, e.g. en / fr / de / nl",
  "english_translation": "English translation (identical to text if already English)",
  "category": "one of: carbon | biodiversity | packaging | water | sourcing | certification | waste | general",
  "confidence": 0.85
}

If the page contains no environmental claims, return [].
No preamble, no markdown fences, no trailing text — just the JSON array."""


# ── Verification (confidence threshold, no second model) ───────────────────

CONFIDENCE_THRESHOLD = 0.8  # claims with confidence >= this are marked verified


def apply_confidence_threshold(claims: list[dict], threshold: float = CONFIDENCE_THRESHOLD,
                                verbose: bool = True) -> list[dict]:
    """
    Flags each claim as verified/not based on the extraction model's own
    "confidence" score, instead of sending it to a second model for an
    independent check. Adds "confidence_verified" (bool) to each claim —
    True if confidence >= threshold, False otherwise. Missing/unparseable
    confidence values are treated as 0 (i.e. not verified).
    """
    if not claims:
        return claims

    for c in claims:
        try:
            conf = float(c.get("confidence", 0))
        except (TypeError, ValueError):
            conf = 0.0
        c["confidence_verified"] = conf >= threshold

    if verbose:
        verified_count = sum(1 for c in claims if c.get("confidence_verified"))
        print(f"[confidence >= {threshold}: {verified_count}/{len(claims)} passed]", end=" ")

    return claims


# ── Core extraction ────────────────────────────────────────────────────────────

def extract_claims_from_page(client, company_name: str, page_name: str, page_url: str, page_text: str) -> list[dict]:
    """Call Claude API for one page and return a list of claim dicts."""
    user_msg = f"Company: {company_name}\nPage: {page_name}\nURL: {page_url}\n\nPage text:\n{page_text[:8000]}"

    try:
        message = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=10000,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": user_msg}],
        )
        raw = message.content[0].text.strip()

        # Strip accidental markdown fences
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
            raw = raw.strip()

        claims = json.loads(raw)
        if not isinstance(claims, list):
            return []

        # Attach page metadata to each claim
        for c in claims:
            c["page"] = page_name
            c["page_url"] = page_url

        return claims

    except json.JSONDecodeError as e:
        # Common cause: a verbatim claim contains an actual quote
        # character (a testimonial, guillemets normalized to straight
        # quotes, etc.) that the model didn't escape, breaking strict
        # JSON syntax. Try a permissive repair before giving up —
        # recovers the claims instead of silently losing the whole page.
        try:
            claims = json_repair.loads(raw)
            if isinstance(claims, list) and claims:
                print(f"    [WARN] JSON parse error on '{page_name}' ({e}) — "
                      f"recovered {len(claims)} claims via json_repair")
                for c in claims:
                    c["page"] = page_name
                    c["page_url"] = page_url
                return claims
        except Exception:
            pass

        print(f"    [WARN] JSON parse error on '{page_name}': {e}")
        # Save the raw response so we can inspect exactly what broke,
        # instead of silently discarding it.
        debug_dir = Path("debug_failed_json")
        debug_dir.mkdir(exist_ok=True)
        debug_path = debug_dir / f"{company_name.lower().replace(' ', '_')}_{page_name.lower()}.txt"
        try:
            with open(debug_path, "w", encoding="utf-8") as f:
                f.write(raw)
            print(f"    [WARN] Raw response saved to: {debug_path}")
        except Exception:
            pass
        return []
    except anthropic.APIError as e:
        print(f"    [ERROR] API error on '{page_name}': {e}")
        return []


def process_scrape_result(anthropic_client, data: dict, source_file: str = "",
                           confidence_threshold: float = CONFIDENCE_THRESHOLD,
                           verbose: bool = True) -> dict:
    """
    Process one company's scrape result — takes the SAME dict shape that
    run_single_scrape() / scrape_sme_website() returns, whether it came
    from a freshly-scraped in-memory result or was loaded from JSON.

    This is the function pipelineV2.py calls directly, so Stage 1 output
    can flow straight into Stage 2 without a mandatory disk round-trip.

    Args:
        anthropic_client: an anthropic.Anthropic client instance (extraction)
        data: dict with keys company_name, base_url, scraped_at, pages
        source_file: optional label for the Excel "Source File" column
        confidence_threshold: claims with confidence >= this are marked verified
        verbose: whether to print progress to stdout

    Returns:
        dict with company_name, base_url, scraped_at, source_file, claims
    """
    company_name = data.get("company_name", "Unknown Company")
    base_url = data.get("base_url", "")
    scraped_at = data.get("scraped_at", "")
    pages = data.get("pages", {})

    if verbose:
        print(f"\n{'─'*60}")
        print(f"  Company : {company_name}")
        if source_file:
            print(f"  File    : {source_file}")
        print(f"  Pages   : {list(pages.keys())}")

    all_claims = []
    for page_name, page_data in pages.items():
        page_text = page_data.get("text", "").strip()
        page_url = page_data.get("url", "")

        if len(page_text) < 50:
            if verbose:
                print(f"    [{page_name}] skipped — too short")
            continue

        if verbose:
            print(f"    [{page_name}] extracting...", end=" ", flush=True)
        claims = extract_claims_from_page(anthropic_client, company_name, page_name, page_url, page_text)
        if verbose:
            print(f"{len(claims)} claims found", end=" ")

        if claims:
            claims = apply_confidence_threshold(claims, threshold=confidence_threshold, verbose=verbose)

        if verbose:
            print()

        all_claims.extend(claims)

    if verbose:
        print(f"  Total   : {len(all_claims)} claims")

    return {
        "company_name": company_name,
        "base_url": base_url,
        "scraped_at": scraped_at,
        "source_file": source_file,
        "claims": all_claims,
    }


def process_company_file(anthropic_client, json_path: Path,
                          confidence_threshold: float = CONFIDENCE_THRESHOLD) -> dict:
    """Process one company JSON file and return extracted claims + metadata.
    Thin wrapper around process_scrape_result() for the file-based CLI/batch path."""
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    return process_scrape_result(anthropic_client, data, source_file=json_path.name,
                                  confidence_threshold=confidence_threshold)


# ── Excel output ───────────────────────────────────────────────────────────────

CATEGORY_COLORS = {
    "carbon":        "D6EAF8",
    "biodiversity":  "D5F5E3",
    "packaging":     "FAD7A0",
    "water":         "AED6F1",
    "sourcing":      "E8DAEF",
    "certification": "FDEBD0",
    "waste":         "D5DBDB",
    "general":       "F9EBEA",
}

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def write_excel(results: list[dict], output_path: Path):
    wb = Workbook()

    # ── Sheet 1: All claims ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "All Claims"

    headers = [
        "Company", "Page", "Category", "Verbatim Claim",
        "English Translation", "Language", "Confidence",
        "Confidence Verified",
        "Page URL", "Source File",
        "EMAS Verified", "EU Organic Verified", "B Corp Verified", "Bord Bia Verified",
        "Biopartenaire Verified", "BioED Verified"
    ]
    col_widths = [22, 16, 14, 60, 60, 10, 12, 16, 45, 25, 14, 16, 14, 15, 18, 14]

    # Header row
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    row_idx = 2
    for res in results:
        for c in res["claims"]:
            conf = c.get("confidence", 0)
            cat = c.get("category", "general").lower()
            fill_color = CATEGORY_COLORS.get(cat, "FFFFFF")
            row_fill = PatternFill("solid", fgColor=fill_color)

            def _yn(key):
                val = c.get(key)
                if val is None:
                    return ""  # cert stage / verification wasn't run or merged for this claim
                return "Yes" if val else "No"

            values = [
                res["company_name"],
                c.get("page", ""),
                c.get("category", "general"),
                c.get("text", ""),
                c.get("english_translation", c.get("text", "")),
                c.get("language", "en").upper(),
                round(conf, 2),
                _yn("confidence_verified"),
                c.get("page_url", ""),
                res["source_file"],
                _yn("emas_verified"),
                _yn("eu_organic_verified"),
                _yn("bcorp_verified"),
                _yn("bordbia_verified"),
                _yn("biopartenaire_verified"),
                _yn("bioed_verified"),
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = row_fill
                cell.border = thin_border()
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if col_idx in (4, 5):  # claim text columns
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

            ws.row_dimensions[row_idx].height = 40
            row_idx += 1

    # ── Sheet 2: Summary per company ──────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    sum_headers = ["Company", "Source File", "Base URL", "Scraped At",
                   "Total Claims", "carbon", "biodiversity", "packaging",
                   "water", "sourcing", "certification", "waste", "general",
                   "Languages Detected"]
    sum_widths =  [22, 25, 40, 22, 14, 10, 14, 12, 8, 10, 15, 8, 10, 25]

    for col_idx, (h, w) in enumerate(zip(sum_headers, sum_widths), 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
    ws2.row_dimensions[1].height = 28
    ws2.freeze_panes = "A2"

    categories = ["carbon", "biodiversity", "packaging", "water",
                  "sourcing", "certification", "waste", "general"]

    for row_idx, res in enumerate(results, 2):
        claims = res["claims"]
        cat_counts = {cat: sum(1 for c in claims if c.get("category", "general") == cat) for cat in categories}
        langs = ", ".join(sorted({c.get("language", "en").upper() for c in claims})) or "EN"

        row_data = [
            res["company_name"], res["source_file"], res["base_url"], res["scraped_at"],
            len(claims),
            *[cat_counts[cat] for cat in categories],
            langs,
        ]
        alt_fill = PatternFill("solid", fgColor="F2F3F4" if row_idx % 2 == 0 else "FFFFFF")
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = alt_fill
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center")

    wb.save(output_path)
    print(f"\n  Saved → {output_path}")


# ── CLI ────────────────────────────────────────────────────────────────────────

def collect_json_files(input_path: str) -> list[Path]:
    p = Path(input_path)
    if p.is_file():
        return [p]
    if p.is_dir():
        files = sorted(p.glob("*.json"))
        if not files:
            print(f"[ERROR] No JSON files found in {p}")
            sys.exit(1)
        return files
    print(f"[ERROR] Path not found: {p}")
    sys.exit(1)


def main():
    parser = argparse.ArgumentParser(description="GreenLens — Environmental Claim Extractor (v2)")
    parser.add_argument("--input",  default='/Users/tunaerdem/Desktop/Scraped Data/Scraped companies/Tuna_extract',       help="Path to a JSON file or folder of JSONs (default: ./data)")
    parser.add_argument("--output", default="",             help="Output Excel path (default: greenlens_claims_<timestamp>.xlsx)")
    parser.add_argument("--threshold", type=float, default=CONFIDENCE_THRESHOLD,
                         help=f"Confidence threshold for verification (default: {CONFIDENCE_THRESHOLD})")
    args = parser.parse_args()

    anthropic_key = os.environ.get("ANTHROPIC_API_KEY")
    if not anthropic_key:
        print("[ERROR] ANTHROPIC_API_KEY environment variable not set.")
        sys.exit(1)

    anthropic_client = anthropic.Anthropic(api_key=anthropic_key)

    json_files = collect_json_files(args.input)
    print(f"Found {len(json_files)} JSON file(s) to process.")

    results = []
    for path in json_files:
        result = process_company_file(anthropic_client, path, confidence_threshold=args.threshold)
        results.append(result)

    if not any(r["claims"] for r in results):
        print("\n[WARN] No claims were extracted across all files.")
        sys.exit(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = Path(args.output) if args.output else Path(f"greenlens_claims_{timestamp}.xlsx")
    write_excel(results, output_path)

    total_claims = sum(len(r["claims"]) for r in results)
    print(f"\n{'='*60}")
    print(f"  Done. {total_claims} claims extracted from {len(json_files)} company file(s).")
    print(f"  Output: {output_path}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()