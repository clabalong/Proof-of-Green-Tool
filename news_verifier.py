"""
news_verifier.py

Checks a company for recent negative environmental/greenwashing-related
news coverage using Google News RSS — lawsuits, fines, greenwashing
accusations, regulatory action. This is intentionally NEGATIVE-SIGNAL
focused: positive press ("won a sustainability award") doesn't tell us
anything about whether a company's CLAIMS are accurate, so it's not
what this check looks for.

IMPORTANT LIMITATIONS (Google News RSS):
- This is an unofficial, undocumented feed — no API key, no published
  SLA, no rate-limit guarantees. Google can change or block it without
  notice. Fine for research use; do not treat as a stable production
  dependency.
- Results are whatever Google News' own ranking/aggregation surfaces
  for the query, not a comprehensive archive search. Coverage and
  recall will vary company to company.
- No full article text — only title, link, publish date, and (when
  present) source name. To read a full article you'd need to follow
  the URL and scrape it separately.
- The `when:Nd` search operator restricts results to roughly the last
  N days, mirroring the same "catch anything breaking right now"
  framing as before — not a historical audit.
- Most small food/beverage SMEs simply won't have any news coverage in
  a given window — an empty result is the EXPECTED outcome for the
  large majority of companies checked, not a sign of a bug.
- Google occasionally 429s or serves a CAPTCHA page to scripted
  requests with no distinguishing User-Agent — a plain `requests.get`
  works most of the time but isn't guaranteed. If you see repeated
  empty/malformed results, check the raw response before assuming
  "no coverage."

Usage:
    from news_verifier import check_news_controversy
    news_result = check_news_controversy(company_name, anthropic_client)

No API key required.
"""

import re
import json
import requests
import xml.etree.ElementTree as ET
from urllib.parse import quote

GOOGLE_NEWS_RSS_URL = "https://news.google.com/rss/search"

# Free/unofficial feed — stay conservative on the lookback window.
DAYS_BACK = 29

# Query focuses on negative/controversy signals specifically — see
# module docstring for why positive press isn't included.
# (Google's search syntax doesn't support NewsAPI-style '*' wildcards,
# so "greenwash*" became the explicit terms below.)
CONTROVERSY_TERMS = (
    '(greenwashing OR "greenwashing" OR "misleading claims" OR '
    '"environmental violation" OR lawsuit OR sued OR fined OR fine OR '
    'scandal OR "false advertising" OR "regulatory action")'
)

REQUEST_HEADERS = {
    # A plain requests default User-Agent is more likely to get an
    # atypical/empty response from Google; a normal browser-ish UA
    # is more reliable for this unofficial endpoint.
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    )
}

VERIFICATION_MODEL = "claude-sonnet-4-6"

NEWS_INTERPRETATION_PROMPT = """You are reviewing recent news search results for a food/beverage company, as part of a greenwashing compliance research project.

Company name: {company_name}

Below are news articles returned by a search for this company name combined with controversy-related terms (lawsuit, fine, greenwashing, etc.). Note: company names can be ambiguous — some results may be about a DIFFERENT company or organization that happens to share the name, or may not be genuinely about environmental/greenwashing matters at all.

Articles (title, description, source, date):
{articles_json}

For each article, determine:
1. Is this genuinely about the SAME company (a food/beverage SME), not a namesake or unrelated entity?
2. Does it describe an actual environmental controversy, greenwashing accusation, lawsuit, fine, or regulatory action — not just a passing/irrelevant mention?

Return ONLY valid JSON, no markdown fences, no other text, in this exact format:
{{
  "controversy_detected": true or false,
  "summary": "1-2 sentence plain-English summary of what was found, or 'No relevant negative coverage found' if nothing qualifies",
  "flagged_articles": [
    {{"title": "...", "url": "...", "reason": "why this is relevant"}}
  ]
}}

If no articles are genuinely relevant, return controversy_detected: false, an appropriate summary, and an empty flagged_articles list."""


def _strip_html(text: str) -> str:
    """Google's RSS <description> field is often an HTML snippet
    (a link plus a font-colored source tag) rather than plain text."""
    text = re.sub(r"<[^>]+>", " ", text or "")
    text = re.sub(r"&nbsp;", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def search_company_news(company_name: str, days_back: int = DAYS_BACK, timeout: int = 15) -> list[dict]:
    """
    Searches Google News RSS for recent articles mentioning the
    company name alongside controversy-related terms.

    Returns a list of {"title", "description", "url", "source", "publishedAt"}
    dicts, or an empty list on failure / no results.
    """
    query = f'"{company_name}" {CONTROVERSY_TERMS} when:{days_back}d'
    params = {
        "q": query,
        "hl": "en-US",
        "gl": "US",
        "ceid": "US:en",
    }

    try:
        resp = requests.get(
            GOOGLE_NEWS_RSS_URL,
            params=params,
            headers=REQUEST_HEADERS,
            timeout=timeout,
        )
        resp.raise_for_status()
    except Exception as e:
        print(f"  [WARN] Google News RSS request failed: {type(e).__name__} — {e}")
        return []

    try:
        root = ET.fromstring(resp.content)
    except ET.ParseError as e:
        print(f"  [WARN] Could not parse Google News RSS response: {e}")
        return []

    articles = []
    for item in root.findall(".//item"):
        title = (item.findtext("title") or "").strip()
        link = (item.findtext("link") or "").strip()
        pub_date = (item.findtext("pubDate") or "").strip()
        description = _strip_html(item.findtext("description"))

        source_el = item.find("source")
        source = (source_el.text or "").strip() if source_el is not None else ""

        articles.append({
            "title": title,
            "description": description,
            "url": link,
            "source": source,
            "publishedAt": pub_date,
        })

    return articles


def check_news_controversy(company_name: str, anthropic_client, days_back: int = DAYS_BACK, verbose: bool = True) -> dict:
    """
    Stage entry point: searches for recent negative news coverage and
    uses Claude to interpret whether any of it is genuinely relevant
    (right company, genuinely about an environmental controversy).

    Args:
        company_name: company to check
        anthropic_client: an anthropic.Anthropic client instance
        days_back: lookback window in days (Google 'when:Nd' operator)
        verbose: whether to print progress to stdout

    Returns:
        dict with "controversy_detected" (bool), "summary" (str),
        "flagged_articles" (list), "articles_found" (int, raw count
        before interpretation)
    """
    if verbose:
        print(f"  Checking recent news for: {company_name}")

    articles = search_company_news(company_name, days_back=days_back)

    empty_result = {
        "controversy_detected": False,
        "summary": "No articles found in the last month" if not articles else "",
        "flagged_articles": [],
        "articles_found": len(articles),
    }

    if not articles:
        if verbose:
            print("    No articles found.")
        return empty_result

    prompt = NEWS_INTERPRETATION_PROMPT.format(
        company_name=company_name,
        articles_json=json.dumps(articles, ensure_ascii=False, indent=2),
    )

    try:
        response = anthropic_client.messages.create(
            model=VERIFICATION_MODEL,
            max_tokens=1500,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = response.content[0].text.strip()
        raw = raw.replace("```json", "").replace("```", "").strip()
        result = json.loads(raw)
        result["articles_found"] = len(articles)

        if verbose:
            status = "CONTROVERSY DETECTED" if result.get("controversy_detected") else "no relevant coverage"
            print(f"    {len(articles)} raw article(s) found -> {status}")

        return result

    except Exception as e:
        if verbose:
            print(f"    [WARN] News interpretation failed: {type(e).__name__} — {e}")
        empty_result["summary"] = f"Interpretation failed: {e}"
        return empty_result


def append_news_sheet(excel_path, results: dict):
    """
    Appends a 'News Check' sheet to an existing xlsx workbook, same
    pattern as cert_verifier_api.append_certifications_sheet().

    Args:
        excel_path: path to the .xlsx file already on disk
        results: dict of {company_name: news_check_result}
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(excel_path)
    ws = wb.create_sheet("News Check")

    headers = ["Company", "Articles Found", "Controversy Detected", "Summary", "Flagged Article", "URL", "Reason"]
    widths = [22, 14, 18, 50, 40, 45, 45]
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"

    row_idx = 2
    for company_name, result in results.items():
        flagged = result.get("flagged_articles", [])
        if not flagged:
            ws.cell(row=row_idx, column=1, value=company_name)
            ws.cell(row=row_idx, column=2, value=result.get("articles_found", 0))
            ws.cell(row=row_idx, column=3, value="Yes" if result.get("controversy_detected") else "No")
            ws.cell(row=row_idx, column=4, value=result.get("summary", ""))
            row_idx += 1
        else:
            for article in flagged:
                ws.cell(row=row_idx, column=1, value=company_name)
                ws.cell(row=row_idx, column=2, value=result.get("articles_found", 0))
                ws.cell(row=row_idx, column=3, value="Yes" if result.get("controversy_detected") else "No")
                ws.cell(row=row_idx, column=4, value=result.get("summary", ""))
                ws.cell(row=row_idx, column=5, value=article.get("title", ""))
                ws.cell(row=row_idx, column=6, value=article.get("url", ""))
                ws.cell(row=row_idx, column=7, value=article.get("reason", ""))
                row_idx += 1

    wb.save(excel_path)


if __name__ == "__main__":
    import os
    import anthropic
    client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))
    for name in ["Danone", "Glenisk"]:
        print(f"\n--- {name} ---")
        result = check_news_controversy(name, client)
        print(json.dumps(result, indent=2, ensure_ascii=False))