"""
cert_verifier_api.py

Unified, API-driven certification checker for EMAS, EU Organic, and B Corp
— replaces static CSV scrapes with live/cached lookups.

Design:
- B Corp (Typesense): TRUE live search. Query the company name directly
  against the index — no pre-fetching needed.
- EMAS & EU Organic: neither API supports name search, only country/status
  filters. So these are "fetch full country list once, cache in memory,
  fuzzy-match locally" — still dynamic in the sense that re-running later
  pulls fresh data, but each check_company() call doesn't hit the network
  per company; it hits the network once per country (cached after that).

Usage:
    from cert_verifier_api import CertVerifier

    v = CertVerifier(countries=["Ireland", "France", "Belgium", "Austria"])
    result = v.check_company("Glenisk")
    print(result)
    # {
    #   "emas": {"matched": False, "best_match": None, "score": 41},
    #   "eu_organic": {"matched": True, "best_match": {...}, "score": 92},
    #   "bcorp": {"matched": False, "best_match": None, "score": 0},
    # }

Pipeline usage (Stage 3, chained after Stage 1/2 via pipelineV1.py):
    from cert_verifier_api import run_certification_stage, append_certifications_sheet

    cert_result = run_certification_stage("Glenisk")
    append_certifications_sheet(excel_path, {"Glenisk": cert_result})

Fuzzy matching uses difflib (stdlib, no extra install). Swap in
rapidfuzz for better results on large datasets if needed — same interface.
"""

import difflib
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup

# ---------------------------------------------------------------------------
# EMAS
# ---------------------------------------------------------------------------
EMAS_URL = "https://webgate.ec.europa.eu/emas2/public/rest/registration/list"

# ---------------------------------------------------------------------------
# EU Organic (TRACES NT)
# ---------------------------------------------------------------------------
ORGANIC_URL = "https://webgate.ec.europa.eu/tracesnt/directory/publication/organic-operator/for/query"
ORGANIC_PAGE_SIZE = 100

# ---------------------------------------------------------------------------
# B Corp (Typesense)
# ---------------------------------------------------------------------------
BCORP_URL = "https://94eo8lmsqa0nd3j5p.a1.typesense.net/multi_search"
BCORP_API_KEY = "eoWf8NTNsTFdaxcxNSuyaKAjLeV4T3F0"
BCORP_COLLECTION = "companies-production-en-us"
BCORP_QUERY_BY = (
    "name,description,websiteKeywords,countries,industry,sector,"
    "hqCountry,hqProvince,hqCity,hqPostalCode,provinces,cities,size,"
    "demographicsList"
)

# Country name -> ISO2, only the ones this project needs; extend as needed
COUNTRY_ISO2 = {
    "Ireland": "IE",
    "France": "FR",
    "Belgium": "BE",
    "Austria": "AT",
}

# Default panel for pipeline use — matches the GreenLens dissertation sample.
DEFAULT_COUNTRIES = ["Ireland", "France", "Belgium", "Austria"]

# Reverse of COUNTRY_ISO2, used to guess a company's country from its
# domain TLD (e.g. glenisk.com won't match, but a hypothetical
# glenisk.ie would guess "Ireland").
TLD_TO_COUNTRY = {v.lower(): k for k, v in COUNTRY_ISO2.items()}


def guess_country_from_url(url: str):
    """
    Best-effort guess of a company's country from its domain's TLD.
    Returns the country name (e.g. "Ireland") if the TLD confidently
    maps to one of the panel countries, or None if it doesn't
    (.com, .nl, .de, generic domains, etc.) — callers should check ALL
    countries in that case, since guessing wrong risks silently
    missing a real registry match.
    """
    from urllib.parse import urlparse
    netloc = urlparse(url if "://" in url else "https://" + url).netloc
    netloc = netloc.replace("www.", "")
    tld = netloc.split(".")[-1].lower()
    return TLD_TO_COUNTRY.get(tld)


# ---------------------------------------------------------------------------
# Bord Bia — Origin Green Members Directory (local cache, not a live API)
# ---------------------------------------------------------------------------
# Origin Green's site has bot-detection that makes live scraping fragile,
# so this uses a pre-scraped local JSON cache instead of a live lookup —
# same "fetch once, match locally" idea as EMAS/EU Organic, just backed
# by a static file instead of a network call.
BORDBIA_CACHE_PATH = "bordbia_members_cache.json"

# Matches a domain anywhere in a string, with or without scheme/www,
# and regardless of surrounding junk text (many cache entries are messy,
# e.g. "http://www.site.com also attended Workshop 07/03/2014").
_DOMAIN_PATTERN = re.compile(
    r"(?:https?://)?(?:www\.)?([a-zA-Z0-9](?:[a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?"
    r"(?:\.[a-zA-Z0-9](?:[a-zA-Z0-9-]{0,61}[a-zA-Z0-9])?)+)",
    re.IGNORECASE,
)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0 Safari/537.36",
}

MATCH_THRESHOLD = 90  # 0-100, similarity score below which "matched" is False
# Raised from 80 after testing: with suffix-stripping + word-boundary
# containment bonus, genuine matches (same company, different legal
# suffix) score 95-100, while coincidental short-name overlaps (e.g.
# "Danival" vs "Aniva SRL", sharing 5 of 7 letters by chance) can still
# score up to ~83 on raw character similarity alone. 90 cleanly separates
# the two based on empirical testing across the SME sample.

# Common legal/corporate suffixes across the target jurisdictions. Stripped
# before comparison so "Glenisk" vs "Glenisk Ltd" scores as a near-perfect
# match instead of being penalized for the extra word.
COMPANY_SUFFIXES = {
    "ltd", "limited", "gmbh", "inc", "incorporated", "plc", "co", "corp",
    "corporation", "llc", "sarl", "sas", "sa", "nv", "bv", "ag", "spa",
    "srl", "kg", "gmbh & co kg", "eurl", "gie", "ug", "cic", "coop",
    "cooperative", "group", "holdings", "holding",
}


def _normalize_name(name: str) -> str:
    """Lowercases, strips punctuation, and drops trailing legal/corporate
    suffix words (Ltd, GmbH, SARL, etc.) so name comparisons aren't
    penalized for boilerplate company-type wording."""
    name = (name or "").lower().strip()
    name = re.sub(r"[^\w\s]", " ", name)
    words = [w for w in name.split() if w]
    while words and words[-1] in COMPANY_SUFFIXES:
        words.pop()
    return " ".join(words)


def _is_word_contained(shorter: str, longer: str) -> bool:
    """
    True only if `shorter` appears in `longer` as a whole word or whole
    phrase — anchored to word boundaries — not as a coincidental substring
    buried inside a longer word (e.g. "aniva" inside "danival" should NOT
    count, but "glenisk" inside "glenisk dairy products" should).
    """
    if not shorter:
        return False
    pattern = r"(?<!\w)" + re.escape(shorter) + r"(?!\w)"
    return re.search(pattern, longer) is not None


def _similarity(a: str, b: str) -> int:
    na, nb = _normalize_name(a), _normalize_name(b)
    ratio = difflib.SequenceMatcher(None, na, nb).ratio()

    # Containment bonus: if one normalized name is fully contained in the
    # other AT WORD BOUNDARIES (e.g. "glenisk" inside "glenisk dairy
    # products"), treat it as a strong match even if the raw
    # character-ratio dips below threshold purely due to length
    # difference. Word-boundary anchoring prevents false positives like
    # "aniva" matching inside "danival" (a coincidental substring, not a
    # shared word).
    if na and nb:
        shorter, longer = (na, nb) if len(na) <= len(nb) else (nb, na)
        if _is_word_contained(shorter, longer):
            ratio = max(ratio, 0.95)

    return round(ratio * 100)


def _best_match(name: str, candidates: list[dict], name_key: str):
    best, best_score = None, 0
    for c in candidates:
        candidate_name = c.get(name_key) or ""
        score = _similarity(name, candidate_name)
        if score > best_score:
            best, best_score = c, score
    return best, best_score


@dataclass
class CertVerifier:
    countries: list[str]  # human-readable, e.g. ["Ireland", "France"]
    _emas_cache: dict = field(default_factory=dict)      # country -> list[dict]
    _organic_cache: dict = field(default_factory=dict)   # country -> list[dict]

    # -- EMAS -----------------------------------------------------------
    def _load_emas(self) -> list[dict]:
        """Fetch once, filter to target countries, cache in memory."""
        if self._emas_cache:
            return [r for recs in self._emas_cache.values() for r in recs]

        resp = requests.get(EMAS_URL, headers=HEADERS, timeout=60)
        resp.raise_for_status()
        all_records = resp.json()

        for country in self.countries:
            self._emas_cache[country] = [
                r for r in all_records if r.get("organisationCountryName") == country
            ]
        return [r for recs in self._emas_cache.values() for r in recs]

    def check_emas(self, company_name: str) -> dict:
        records = self._load_emas()
        best, score = _best_match(company_name, records, "organisationName")
        return {"matched": score >= MATCH_THRESHOLD, "best_match": best, "score": score}

    # -- EU Organic -------------------------------------------------------
    def _load_organic_country(self, country: str) -> list[dict]:
        if country in self._organic_cache:
            return self._organic_cache[country]

        code = COUNTRY_ISO2.get(country)
        if not code:
            self._organic_cache[country] = []
            return []

        records, offset = [], 0
        while True:
            params = {
                "countryCode": code, "max": ORGANIC_PAGE_SIZE, "offset": offset,
                "sort": "-issuedOn", "states": "ISSUED",
            }
            resp = requests.get(ORGANIC_URL, headers=HEADERS, params=params, timeout=30)
            if resp.status_code == 500:
                break  # hit the ~10k pagination cap; good enough for cached lookups
            resp.raise_for_status()
            page = resp.json()
            if not page:
                break
            records.extend(page)
            if len(page) < ORGANIC_PAGE_SIZE:
                break
            offset += ORGANIC_PAGE_SIZE

        self._organic_cache[country] = records
        return records

    def check_eu_organic(self, company_name: str) -> dict:
        all_records = []
        for country in self.countries:
            all_records.extend(self._load_organic_country(country))

        # operator name is nested: record["operator"]["name"]
        flat = [{"name": (r.get("operator") or {}).get("name", ""), "_raw": r} for r in all_records]
        best, score = _best_match(company_name, flat, "name")
        return {
            "matched": score >= MATCH_THRESHOLD,
            "best_match": best["_raw"] if best else None,
            "score": score,
        }

    # -- B Corp (true live search) ---------------------------------------
    def check_bcorp(self, company_name: str) -> dict:
        payload = {
            "searches": [{
                "collection": BCORP_COLLECTION,
                "q": company_name,
                "query_by": BCORP_QUERY_BY,
                "per_page": 5,
            }]
        }
        resp = requests.post(
            BCORP_URL, headers={**HEADERS, "Content-Type": "application/json"},
            params={"x-typesense-api-key": BCORP_API_KEY}, json=payload, timeout=15,
        )
        resp.raise_for_status()
        hits = resp.json()["results"][0].get("hits", [])
        if not hits:
            return {"matched": False, "best_match": None, "score": 0}

        docs = [h.get("document", h) for h in hits]
        best, score = _best_match(company_name, docs, "name")
        return {"matched": score >= MATCH_THRESHOLD, "best_match": best, "score": score}

    # -- Combined ----------------------------------------------------------
    def check_company(self, company_name: str) -> dict:
        return {
            "emas": self.check_emas(company_name),
            "eu_organic": self.check_eu_organic(company_name),
            "bcorp": self.check_bcorp(company_name),
        }


# ================================================================
# BORD BIA — ORIGIN GREEN (local cache lookup)
# ================================================================

# Module-level cache keyed by resolved cache path, so repeated calls
# (e.g. across a batch of companies) only read/parse the JSON file once.
_bordbia_cache_store: dict = {}


def _extract_domain(text: str):
    """Finds a domain anywhere in a string, or None if none present."""
    if not text:
        return None
    match = _DOMAIN_PATTERN.search(text)
    if not match:
        return None
    domain = match.group(1).lower()
    if domain.startswith("www."):
        domain = domain[4:]
    return domain if "." in domain else None


def _load_bordbia_cache(cache_path: str = None):
    """
    Loads and indexes the Bord Bia members cache once, caching the
    result in memory for subsequent calls.

    Returns (domain_index, name_candidates):
        domain_index: dict of {normalized_domain: raw_display_string}
        name_candidates: list of {"name": ..., "_raw": ...} for fuzzy
                          name matching fallback
    """
    resolved_path = cache_path or BORDBIA_CACHE_PATH
    if resolved_path in _bordbia_cache_store:
        return _bordbia_cache_store[resolved_path]

    path = Path(resolved_path)
    if not path.exists():
        print(f"  [WARN] Bord Bia cache not found at '{resolved_path}' — "
              f"skipping Bord Bia check.")
        empty = ({}, [])
        _bordbia_cache_store[resolved_path] = empty
        return empty

    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    members = data.get("members_detailed", [])

    domain_index = {}
    name_candidates = []
    for m in members:
        search_name = m.get("search_name", "")
        display = m.get("display", search_name)

        domain = _extract_domain(search_name) or _extract_domain(display)
        if domain:
            domain_index[domain] = display

        # Also keep every entry as a name-matching fallback candidate —
        # harmless for messy/domain entries since they'll simply score
        # low against real company names and never pass the threshold.
        name_candidates.append({"name": search_name, "_raw": display})

    result = (domain_index, name_candidates)
    _bordbia_cache_store[resolved_path] = result
    return result


def check_bordbia(company_name: str, company_url: str = None, cache_path: str = None) -> dict:
    """
    Checks whether a company appears in the Bord Bia Origin Green
    members directory cache.

    Matching strategy, in priority order:
      1. Domain match — if company_url is given, its domain is compared
         directly against domains found in the cache (most entries are
         raw URLs, not clean names, so this is the primary signal and
         is treated as authoritative: exact domain match -> score 100).
      2. Fuzzy name match — fallback using the same normalization/
         containment logic as check_emas/check_eu_organic/check_bcorp,
         for the minority of cache entries that have a clean company name.

    Args:
        company_name: company name to check
        company_url: company's website URL (recommended — most cache
                      entries can only be matched by domain, not name)
        cache_path: path to bordbia_members_cache.json, defaults to
                     BORDBIA_CACHE_PATH in the current working directory

    Returns:
        dict with "matched", "best_match" (raw display string or None),
        "score", and "match_type" ("domain" | "name" | None)
    """
    domain_index, name_candidates = _load_bordbia_cache(cache_path)

    if company_url:
        company_domain = _extract_domain(company_url)
        if company_domain and company_domain in domain_index:
            return {
                "matched": True,
                "best_match": domain_index[company_domain],
                "score": 100,
                "match_type": "domain",
            }

    if name_candidates:
        best, score = _best_match(company_name, name_candidates, "name")
        return {
            "matched": score >= MATCH_THRESHOLD,
            "best_match": best["_raw"] if best else None,
            "score": score,
            "match_type": "name" if best else None,
        }

    return {"matched": False, "best_match": None, "score": 0, "match_type": None}


# ================================================================
# BIOPARTENAIRE & BIOED — live HTTP checks (no Playwright needed)
# ================================================================
# Both member directories are server-rendered plain HTML (confirmed by
# testing), so unlike Bord Bia these can be fetched live with a normal
# GET request — genuinely dynamic, same idea as B Corp's live search,
# just without a real search API. Cached in memory per session so a
# batch run across many companies only fetches each page once.

BIOPARTENAIRE_URL = "https://www.biopartenaire.com/fr/qui-sommes-nous/"
BIOED_URL = "https://bioed.fr/la-communaute/"

# Domains that show up as links on these pages but are never a member's
# own site (social media, the label's own domain, etc.) — excluded so
# they never get treated as a "member domain".
_NON_MEMBER_DOMAINS = (
    "biopartenaire.com", "bioed.fr", "synabio.com", "cosmebio.org",
    "facebook.com", "instagram.com", "youtube.com", "linkedin.com",
    "twitter.com", "x.com",
)

_biopartenaire_cache = None  # None until first fetch; then dict {domain: raw_href}
_bioed_cache = None  # None until first fetch; then list of member name strings


def _load_biopartenaire():
    global _biopartenaire_cache
    if _biopartenaire_cache is not None:
        return _biopartenaire_cache

    try:
        resp = requests.get(BIOPARTENAIRE_URL, headers=HEADERS, timeout=20)
        resp.raise_for_status()
    except Exception as e:
        print(f"  [WARN] Could not fetch Biopartenaire member list: {e}")
        _biopartenaire_cache = {}
        return _biopartenaire_cache

    soup = BeautifulSoup(resp.text, "html.parser")
    domain_index = {}
    for a in soup.find_all("a", href=True):
        href = a["href"]
        domain = _extract_domain(href)
        if not domain or any(nd in domain for nd in _NON_MEMBER_DOMAINS):
            continue
        domain_index[domain] = href

    _biopartenaire_cache = domain_index
    return domain_index


def check_biopartenaire(company_name: str, company_url: str = None) -> dict:
    """
    Checks Biopartenaire (French fair-trade organic label) membership.
    Domain-matching only for now — the page is a logo grid linking
    directly to each member's own website, which is the reliable signal
    here (no clean company-name text is exposed per logo).

    Args:
        company_name: kept for interface consistency with other checks
        company_url: the company's website URL — REQUIRED for a match,
                      since this check has no name-based fallback

    Returns:
        dict with "matched", "best_match" (raw href or None), "score",
        "match_type" ("domain" | None)
    """
    domain_index = _load_biopartenaire()
    if company_url:
        company_domain = _extract_domain(company_url)
        if company_domain and company_domain in domain_index:
            return {
                "matched": True,
                "best_match": domain_index[company_domain],
                "score": 100,
                "match_type": "domain",
            }
    return {"matched": False, "best_match": None, "score": 0, "match_type": None}


def _load_bioed():
    global _bioed_cache
    if _bioed_cache is not None:
        return _bioed_cache

    try:
        resp = requests.get(BIOED_URL, headers=HEADERS, timeout=20)
        resp.raise_for_status()
    except Exception as e:
        print(f"  [WARN] Could not fetch BioED member list: {e}")
        _bioed_cache = []
        return _bioed_cache

    soup = BeautifulSoup(resp.text, "html.parser")
    names = [tag.get_text(strip=True) for tag in soup.find_all("span", class_="title")]
    _bioed_cache = [n for n in names if n]
    return _bioed_cache


def check_bioed(company_name: str) -> dict:
    """
    Checks BioED (BioEntrepriseDurable, French CSR label) membership.
    Name-based fuzzy matching only — the page doesn't expose member
    website links, just clean company names.

    Returns:
        dict with "matched", "best_match" (raw member name or None),
        "score", "match_type" ("name" | None)
    """
    names = _load_bioed()
    if not names:
        return {"matched": False, "best_match": None, "score": 0, "match_type": None}

    candidates = [{"name": n, "_raw": n} for n in names]
    best, score = _best_match(company_name, candidates, "name")
    return {
        "matched": score >= MATCH_THRESHOLD,
        "best_match": best["_raw"] if best else None,
        "score": score,
        "match_type": "name" if best else None,
    }


# ================================================================
# PIPELINE STAGE WRAPPER  (Stage 3)
# ================================================================

def run_certification_stage(company_name: str, countries: list[str] = None,
                             company_url: str = None, bordbia_cache_path: str = None,
                             verifier: "CertVerifier" = None, verbose: bool = True) -> dict:
    """
    Stage 3 entry point for pipelineV1.py. Checks one company across
    EMAS / EU Organic / B Corp / Bord Bia (Origin Green).

    Args:
        company_name: the company name to check (e.g. from scrape_result["company_name"])
        countries: explicit list of countries to check. Takes priority over
                    company_url guessing if both are given.
        company_url: the company's website URL. If given and no explicit
                      `countries` list is passed, the TLD is used to guess
                      a single country (e.g. glenisk.ie -> Ireland), which
                      speeds up the EU Organic check specifically (it's
                      the only registry that does per-country paginated
                      fetching — EMAS downloads its full dataset regardless,
                      and B Corp doesn't use a country filter at all). Falls
                      back to DEFAULT_COUNTRIES if the TLD isn't recognized,
                      so a wrong or ambiguous domain never silently causes
                      a missed match. This URL is also the PRIMARY signal
                      for the Bord Bia check, since most of that cache is
                      indexed by domain rather than clean company name.
        bordbia_cache_path: path to bordbia_members_cache.json. Defaults
                             to BORDBIA_CACHE_PATH if not given.
        verifier: an existing CertVerifier instance to reuse (recommended
                   when checking multiple companies back-to-back, so the
                   EMAS/EU Organic country data is only fetched once
                   instead of once per company). Creates a new one if
                   not given.
        verbose: whether to print a short summary to stdout.

    Returns:
        dict with keys "emas", "eu_organic", "bcorp", "bordbia".
    """
    if verifier is None:
        if countries is None and company_url:
            guessed = guess_country_from_url(company_url)
            if guessed:
                countries = [guessed]
                if verbose:
                    print(f"  Guessed country from URL: {guessed} (checking this country only)")
            elif verbose:
                print(f"  Could not guess country from URL — checking all of {DEFAULT_COUNTRIES}")
        verifier = CertVerifier(countries=countries or DEFAULT_COUNTRIES)

    if verbose:
        print(f"  Checking certifications for: {company_name}")

    result = verifier.check_company(company_name)
    result["bordbia"] = check_bordbia(company_name, company_url=company_url, cache_path=bordbia_cache_path)
    result["biopartenaire"] = check_biopartenaire(company_name, company_url=company_url)
    result["bioed"] = check_bioed(company_name)

    if verbose:
        for registry, info in result.items():
            status = "MATCH" if info["matched"] else "no match"
            print(f"    {registry:<12} {status} (score {info['score']})")

    return result


def append_certifications_sheet(excel_path, results: dict):
    """
    Appends a 'Certifications' sheet to an EXISTING xlsx workbook
    (e.g. the one Stage 2's write_excel() already created), without
    touching extract_claims.py.

    Args:
        excel_path: path to the .xlsx file already on disk
        results: dict of {company_name: cert_check_dict}, where each
                 cert_check_dict is the return value of check_company()
                 / run_certification_stage()
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = load_workbook(excel_path)
    ws = wb.create_sheet("Certifications")

    headers = ["Company", "Registry", "Matched", "Score", "Matched Name"]
    widths = [22, 16, 10, 8, 40]
    header_fill = PatternFill("solid", fgColor="2C3E50")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_idx, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A2"

    row_idx = 2
    for company_name, cert_result in results.items():
        for registry, info in cert_result.items():
            bm = info.get("best_match")
            matched_name = None
            if bm:
                if isinstance(bm, str):
                    matched_name = bm  # Bord Bia returns a raw display string
                else:
                    matched_name = (
                        bm.get("organisationName")
                        or bm.get("name")
                        or (bm.get("operator") or {}).get("name")
                    )
            ws.cell(row=row_idx, column=1, value=company_name)
            ws.cell(row=row_idx, column=2, value=registry)
            ws.cell(row=row_idx, column=3, value="Yes" if info.get("matched") else "No")
            ws.cell(row=row_idx, column=4, value=info.get("score", 0))
            ws.cell(row=row_idx, column=5, value=matched_name or "")
            row_idx += 1

    wb.save(excel_path)


def merge_certifications_into_claims(claims_result: dict, cert_result: dict) -> dict:
    """
    Attaches certification verification status directly onto EVERY claim
    for a company, so the flat claims dataset is self-contained for the
    Stage 5 ECGT rules engine — no need to join against a separate sheet.

    Also keeps the full cert_result available at claims_result["certifications"]
    for reference / the separate "Certifications" sheet.

    This matters specifically for claims in the "certification" category
    (e.g. "B Corp certified", "Origin Green member") — the rules engine can
    now directly check whether that claim is independently VERIFIED true,
    which is central to flagging unsubstantiated certification claims
    under Article 6a/6b of Directive 2024/825.

    Args:
        claims_result: the dict returned by process_scrape_result() (Stage 2)
        cert_result: the dict returned by run_certification_stage() (Stage 3)

    Returns:
        claims_result, mutated in place and also returned for convenience
    """
    claims_result["certifications"] = cert_result

    emas_ok = cert_result.get("emas", {}).get("matched", False)
    organic_ok = cert_result.get("eu_organic", {}).get("matched", False)
    bcorp_ok = cert_result.get("bcorp", {}).get("matched", False)
    bordbia_ok = cert_result.get("bordbia", {}).get("matched", False)
    biopartenaire_ok = cert_result.get("biopartenaire", {}).get("matched", False)
    bioed_ok = cert_result.get("bioed", {}).get("matched", False)
    any_ok = emas_ok or organic_ok or bcorp_ok or bordbia_ok or biopartenaire_ok or bioed_ok

    for claim in claims_result.get("claims", []):
        claim["emas_verified"] = emas_ok
        claim["eu_organic_verified"] = organic_ok
        claim["bcorp_verified"] = bcorp_ok
        claim["bordbia_verified"] = bordbia_ok
        claim["biopartenaire_verified"] = biopartenaire_ok
        claim["bioed_verified"] = bioed_ok
        claim["any_certification_verified"] = any_ok

    return claims_result


if __name__ == "__main__":
    verifier = CertVerifier(countries=["Ireland", "France", "Belgium", "Austria"])
    for name in ["Glenisk", "Danival", "Biologon"]:
        print(f"\n--- {name} ---")
        result = verifier.check_company(name)
        for registry, info in result.items():
            status = "MATCH" if info["matched"] else "no match"
            match_name = None
            if info["best_match"]:
                bm = info["best_match"]
                match_name = bm.get("organisationName") or bm.get("name") or (
                    (bm.get("operator") or {}).get("name")
                )
            print(f"  {registry}: {status} (score {info['score']}) -> {match_name}")